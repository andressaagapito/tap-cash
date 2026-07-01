import io
import re
from datetime import datetime, date, timezone
from decimal import Decimal
from typing import Dict, List, Any
from uuid import UUID
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.user import User
from app.models.card import Card
from app.models.expense import Expense, ExpenseType, ExpenseStatus, PaymentMethod
from app.models.expense_installment_payment import ExpenseInstallmentPayment
from app.models.financial_profile import UserFinancialProfile
from app.models.goal import Goal, GoalOption, GoalPriority, GoalCategory, GoalStatus, GoalOptionStatus
from app.models.user_category import UserCategory
from openpyxl.worksheet.datavalidation import DataValidation

# Map Portuguese/English sheet names to internal database keys
SHEET_MAPPING = {
    "usuário": "users",
    "usuario": "users",
    "usuários": "users",
    "usuarios": "users",
    "perfil financeiro": "user_financial_profiles",
    "categorias": "user_categories",
    "cartões": "cards",
    "cartoes": "cards",
    "cartões disponíveis": "cards",
    "cartoes disponiveis": "cards",
    "metas": "goals",
    "opções de metas": "goal_options",
    "opcoes de metas": "goal_options",
    "despesas": "expenses",
    "histórico de parcelas": "expense_installment_payments",
    "historico de parcelas": "expense_installment_payments",
    "pagamento de parcelas": "expense_installment_payments",
    "pagamentos de parcelas": "expense_installment_payments",
    "users": "users",
    "user_financial_profiles": "user_financial_profiles",
    "user_categories": "user_categories",
    "cards": "cards",
    "goals": "goals",
    "goal_options": "goal_options",
    "expenses": "expenses",
    "expense_installment_payments": "expense_installment_payments",
}

# Map column headers to internal database model fields
COLUMN_MAPPING = {
    "nome": "name",
    "name": "name",
    "sobrenome": "last_name",
    "last_name": "last_name",
    "e-mail": "email",
    "email": "email",
    "salário_mensal": "monthly_salary",
    "salário mensal": "monthly_salary",
    "salario mensal": "monthly_salary",
    "monthly_salary": "monthly_salary",
    "marcar_parcelas_pagas_automaticamente": "auto_mark_installments_paid",
    "marcar parcelas pagas automaticamente": "auto_mark_installments_paid",
    "auto_mark_installments_paid": "auto_mark_installments_paid",
    "instituição": "institution",
    "instituicao": "institution",
    "institution": "institution",
    "limite": "limit",
    "limit": "limit",
    "dia_fechamento": "closing_day",
    "dia fechamento": "closing_day",
    "closing_day": "closing_day",
    "dia_vencimento": "due_day",
    "dia vencimento": "due_day",
    "due_day": "due_day",
    "cor": "color",
    "color": "color",
    "ícone": "icon",
    "icone": "icon",
    "icon": "icon",
    "descrição": "description",
    "descricao": "description",
    "description": "description",
    "valor_objetivo": "target_amount",
    "valor objetivo": "target_amount",
    "valor_total_desejado": "target_amount",
    "valor total desejado": "target_amount",
    "target_amount": "target_amount",
    "valor_guardado": "saved_amount",
    "valor guardado": "saved_amount",
    "saved_amount": "saved_amount",
    "prazo": "deadline",
    "deadline": "deadline",
    "prioridade": "priority",
    "priority": "priority",
    "categoria": "category",
    "category": "category",
    "meta": "goal_name",
    "nome_da_meta": "goal_name",
    "nome da meta": "goal_name",
    "goal_id": "goal_id",
    "valor_estimado": "estimated_amount",
    "valor estimado": "estimated_amount",
    "estimated_amount": "estimated_amount",
    "link_referencia": "reference_link",
    "link de referencia": "reference_link",
    "link referencia": "reference_link",
    "link referência": "reference_link",
    "reference_link": "reference_link",
    "cartão": "card_name",
    "cartao": "card_name",
    "card_id": "card_id",
    "método_pagamento": "payment_method",
    "metodo_pagamento": "payment_method",
    "metodo de pagamento": "payment_method",
    "método pagamento": "payment_method",
    "metodo pagamento": "payment_method",
    "forma_pagamento": "payment_method",
    "forma de pagamento": "payment_method",
    "forma pagamento": "payment_method",
    "payment_method": "payment_method",
    "tipo": "type",
    "type": "type",
    "data_compra": "purchase_date",
    "data da compra": "purchase_date",
    "data compra": "purchase_date",
    "purchase_date": "purchase_date",
    "valor_total": "total_amount",
    "valor total": "total_amount",
    "total_amount": "total_amount",
    "valor_parcela": "installment_amount",
    "valor da parcela": "installment_amount",
    "valor parcela": "installment_amount",
    "installment_amount": "installment_amount",
    "total_parcelas": "total_installments",
    "total de parcelas": "total_installments",
    "total parcelas": "total_installments",
    "total_installments": "total_installments",
    "meses_recorrência": "recurrence_months",
    "meses de recorrencia": "recurrence_months",
    "meses recorrencia": "recurrence_months",
    "meses recorrência": "recurrence_months",
    "recurrence_months": "recurrence_months",
    "observações": "notes",
    "observacoes": "notes",
    "notes": "notes",
    "despesa": "expense_name",
    "nome_da_despesa": "expense_name",
    "nome da despesa": "expense_name",
    "expense_id": "expense_id",
    "número_parcela": "installment_number",
    "numero_parcela": "installment_number",
    "numero da parcela": "installment_number",
    "número parcela": "installment_number",
    "numero parcela": "installment_number",
    "installment_number": "installment_number",
    "pago_em": "paid_at",
    "data_pagamento": "paid_at",
    "pago em": "paid_at",
    "paid_at": "paid_at",
    "status": "status",
    "parcelas pagas": "initial_paid_installments",
    "paid_installments": "initial_paid_installments",
}

# Enum Mappings for input processing
PAYMENT_METHOD_MAP = {
    "cartão de crédito": PaymentMethod.CREDIT_CARD,
    "cartao de credito": PaymentMethod.CREDIT_CARD,
    "cartao_credito": PaymentMethod.CREDIT_CARD,
    "credit_card": PaymentMethod.CREDIT_CARD,
    "cartão de débito": PaymentMethod.DEBIT_CARD,
    "cartao de debito": PaymentMethod.DEBIT_CARD,
    "cartao_debito": PaymentMethod.DEBIT_CARD,
    "debit_card": PaymentMethod.DEBIT_CARD,
    "pix": PaymentMethod.PIX,
    "boleto": PaymentMethod.BANK_SLIP,
    "bank_slip": PaymentMethod.BANK_SLIP
}

TYPE_MAP = {
    "única": ExpenseType.ONE_TIME,
    "unica": ExpenseType.ONE_TIME,
    "avulsa": ExpenseType.ONE_TIME,
    "one_time": ExpenseType.ONE_TIME,
    "recorrente": ExpenseType.RECURRING,
    "recurring": ExpenseType.RECURRING
}

STATUS_MAP = {
    "ativo": ExpenseStatus.ACTIVE,
    "ativa": ExpenseStatus.ACTIVE,
    "active": ExpenseStatus.ACTIVE,
    "pago": ExpenseStatus.PAID_OFF,
    "quitada": ExpenseStatus.PAID_OFF,
    "paid_off": ExpenseStatus.PAID_OFF,
    "concluída": ExpenseStatus.PAID_OFF,
    "concluida": ExpenseStatus.PAID_OFF,
    "cancelado": ExpenseStatus.CANCELLED,
    "cancelada": ExpenseStatus.CANCELLED,
    "cancelled": ExpenseStatus.CANCELLED
}

CATEGORY_MAP = {
    "alimentação": "food",
    "alimentacao": "food",
    "food": "food",
    "moradia": "housing",
    "housing": "housing",
    "transporte": "transport",
    "transport": "transport",
    "saúde": "health",
    "saude": "health",
    "health": "health",
    "educação": "education",
    "educacao": "education",
    "education": "education",
    "lazer": "leisure",
    "leisure": "leisure",
    "vestuário": "clothing",
    "vestuario": "clothing",
    "clothing": "clothing",
    "assinaturas": "subscriptions",
    "subscriptions": "subscriptions",
    "outros": "other",
    "outro": "other",
    "other": "other"
}


# Helper to format values for Excel cells
def format_cell_value(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, UUID):
        return str(val)
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, bool):
        return val
    if hasattr(val, "value"):  # Enums
        return val.value
    return val


# Helper to parse decimal values safely
def parse_decimal(val: Any) -> Decimal | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        # Remove any currency symbols, commas or extra whitespace
        cleaned = re.sub(r"[^\d.-]", "", str(val))
        return Decimal(cleaned)
    except Exception:
        return None

# Helper to parse dates safely
def parse_date(val: Any) -> date | None:
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        dt = datetime.fromisoformat(str(val).strip())
        return dt.date()
    except Exception:
        # Fallback to general parsing or date format
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(str(val).strip(), fmt).date()
            except ValueError:
                continue
        return None

# Helper to parse datetimes safely
def parse_datetime(val: Any) -> datetime | None:
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val).strip())
    except Exception:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(str(val).strip(), fmt)
            except ValueError:
                continue
        return None

# Helper to parse boolean safely
def parse_boolean(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "1", "yes", "sim", "t", "y")

# Helper to parse int safely
def parse_int(val: Any) -> int | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        return int(float(str(val).strip()))
    except Exception:
        return None

def generate_excel(user: User, db: Session) -> io.BytesIO:
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # slate-800
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    def write_sheet(title: str, headers: List[str], rows: List[List[Any]]):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_left
            cell.border = thin_border
            
        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=format_cell_value(value))
                cell.font = data_font
                cell.border = thin_border
                
                # Alignments based on value types
                if isinstance(value, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(value, (datetime, date, bool)):
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 1. Usuário
    user_headers = ["Nome", "Sobrenome", "E-mail"]
    user_rows = [[user.name, user.last_name, user.email]]
    write_sheet("Usuário", user_headers, user_rows)

    # 2. Perfil Financeiro
    profile = db.query(UserFinancialProfile).filter(UserFinancialProfile.user_id == user.id).first()
    profile_headers = ["Salário Mensal", "Marcar parcelas pagas automaticamente"]
    profile_rows = []
    if profile:
        profile_rows.append([profile.monthly_salary, "Sim" if profile.auto_mark_installments_paid else "Não"])
    write_sheet("Perfil Financeiro", profile_headers, profile_rows)

    # 3. Categorias
    categories = db.query(UserCategory).filter(UserCategory.user_id == user.id).all()
    categories_headers = ["Nome"]
    categories_rows = [[c.name] for c in categories]
    write_sheet("Categorias", categories_headers, categories_rows)

    # 4. Cartões
    cards = db.query(Card).filter(Card.user_id == user.id).all()
    cards_headers = ["Nome", "Instituição", "Limite", "Dia Fechamento", "Dia Vencimento", "Cor", "Ícone"]
    card_id_to_name = {c.id: c.name for c in cards}
    cards_rows = [
        [c.name, c.institution, c.limit, c.closing_day, c.due_day, c.color, c.icon]
        for c in cards
    ]
    write_sheet("Cartões", cards_headers, cards_rows)

    # 5. Metas
    goals = db.query(Goal).filter(Goal.user_id == user.id).all()
    goals_headers = ["Nome", "Descrição", "Valor Objetivo", "Valor Guardado", "Prazo", "Prioridade", "Categoria", "Status"]
    goal_id_to_name = {g.id: g.name for g in goals}
    
    priority_output_map = {"low": "Baixa", "medium": "Média", "high": "Alta"}
    goal_cat_output_map = {
        "car": "Carro", "travel": "Viagem", "reserve": "Reserva", 
        "debt": "Dívida", "house": "Casa", "education": "Educação", 
        "health": "Saúde", "other": "Outro"
    }
    goal_status_output_map = {"active": "Ativa", "completed": "Concluída", "cancelled": "Cancelada"}
    
    goals_rows = [
        [
            g.name, g.description, g.target_amount, g.saved_amount, g.deadline,
            priority_output_map.get(g.priority.value if hasattr(g.priority, "value") else g.priority, "Média"),
            goal_cat_output_map.get(g.category.value if hasattr(g.category, "value") else g.category, "Outro"),
            goal_status_output_map.get(g.status.value if hasattr(g.status, "value") else g.status, "Ativa")
        ]
        for g in goals
    ]
    write_sheet("Metas", goals_headers, goals_rows)

    # 6. Opções de Metas
    goal_ids = [g.id for g in goals]
    goal_options = []
    if goal_ids:
        goal_options = db.query(GoalOption).filter(GoalOption.goal_id.in_(goal_ids)).all()
    goal_options_headers = ["Meta", "Nome", "Valor Estimado", "Descrição", "Link Referência", "Status"]
    
    option_status_output_map = {"analyzing": "Em análise", "chosen": "Escolhida", "discarded": "Descartada"}
    
    goal_options_rows = [
        [
            goal_id_to_name.get(o.goal_id, ""), o.name, o.estimated_amount, o.description, o.reference_link,
            option_status_output_map.get(o.status.value if hasattr(o.status, "value") else o.status, "Em análise")
        ]
        for o in goal_options
    ]
    write_sheet("Opções de Metas", goal_options_headers, goal_options_rows)

    # 7. Despesas
    expenses = db.query(Expense).filter(Expense.user_id == user.id).all()
    expenses_headers = [
        "Cartão", "Método Pagamento", "Nome", "Tipo", "Data Compra",
        "Valor Total", "Valor Parcela", "Total Parcelas", "Meses Recorrência",
        "Parcelas Pagas", "Categoria", "Observações", "Status"
    ]
    expense_id_to_name = {e.id: e.name for e in expenses}
    
    pm_output_map = {
        PaymentMethod.CREDIT_CARD: "Cartão de crédito",
        PaymentMethod.DEBIT_CARD: "Cartão de débito",
        PaymentMethod.PIX: "Pix",
        PaymentMethod.BANK_SLIP: "Boleto",
    }
    type_output_map = {
        ExpenseType.ONE_TIME: "Única",
        ExpenseType.RECURRING: "Recorrente",
    }
    status_output_map = {
        ExpenseStatus.ACTIVE: "Ativo",
        ExpenseStatus.PAID_OFF: "Pago",
        ExpenseStatus.CANCELLED: "Cancelado",
    }
    category_output_map = {
        "food": "Alimentação",
        "housing": "Moradia",
        "transport": "Transporte",
        "health": "Saúde",
        "education": "Educação",
        "leisure": "Lazer",
        "clothing": "Vestuário",
        "subscriptions": "Assinaturas",
        "other": "Outros",
    }
    
    expenses_rows = [
        [
            card_id_to_name.get(e.card_id, ""),
            pm_output_map.get(e.payment_method, "Cartão de crédito"),
            e.name,
            type_output_map.get(e.type, "Única"),
            e.purchase_date,
            e.total_amount,
            e.installment_amount,
            e.total_installments,
            e.recurrence_months,
            len(e.installment_payments) if e.installment_payments else 0,
            category_output_map.get(e.category, e.category),
            e.notes,
            status_output_map.get(e.status, "Ativo")
        ]
        for e in expenses
    ]
    write_sheet("Despesas", expenses_headers, expenses_rows)

    # 8. Histórico de Parcelas
    expense_ids = [e.id for e in expenses]
    installment_payments = []
    if expense_ids:
        installment_payments = db.query(ExpenseInstallmentPayment).filter(
            ExpenseInstallmentPayment.expense_id.in_(expense_ids)
        ).all()
    payments_headers = ["Despesa", "Número Parcela", "Pago Em"]
    payments_rows = [
        [expense_id_to_name.get(p.expense_id, ""), p.installment_number, p.paid_at]
        for p in installment_payments
    ]
    write_sheet("Histórico de Parcelas", payments_headers, payments_rows)

    # Add Dropdown menus (Data Validation) in Excel
    if "Opções de Metas" in wb.sheetnames and "Metas" in wb.sheetnames:
        ws_opt = wb["Opções de Metas"]
        dv_goals = DataValidation(type="list", formula1="=Metas!$A$2:$A$100", allow_blank=True)
        ws_opt.add_data_validation(dv_goals)
        dv_goals.add("A2:A500")

    if "Despesas" in wb.sheetnames:
        ws_exp = wb["Despesas"]
        if "Cartões" in wb.sheetnames:
            dv_cards = DataValidation(type="list", formula1="=Cartões!$A$2:$A$100", allow_blank=True)
            ws_exp.add_data_validation(dv_cards)
            dv_cards.add("A2:A1000")
            
        dv_pm = DataValidation(type="list", formula1='"Cartão de crédito,Cartão de débito,Pix,Boleto"', allow_blank=True)
        ws_exp.add_data_validation(dv_pm)
        dv_pm.add("B2:B1000")
        
        dv_type = DataValidation(type="list", formula1='"Única,Recorrente"', allow_blank=True)
        ws_exp.add_data_validation(dv_type)
        dv_type.add("D2:D1000")
        
        dv_cat = DataValidation(type="list", formula1='"Alimentação,Moradia,Transporte,Saúde,Educação,Lazer,Vestuário,Assinaturas,Outros"', allow_blank=True)
        ws_exp.add_data_validation(dv_cat)
        dv_cat.add("K2:K1000")
        
        dv_status = DataValidation(type="list", formula1='"Ativo,Pago,Cancelado"', allow_blank=True)
        ws_exp.add_data_validation(dv_status)
        dv_status.add("M2:M1000")

    if "Histórico de Parcelas" in wb.sheetnames and "Despesas" in wb.sheetnames:
        ws_pay = wb["Histórico de Parcelas"]
        dv_expenses = DataValidation(type="list", formula1="=Despesas!$C$2:$C$1000", allow_blank=True)
        ws_pay.add_data_validation(dv_expenses)
        dv_expenses.add("A2:A1000")


def process_excel_upload(file_contents: bytes, user: User, db: Session) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_contents), data_only=True)
    
    summary = {
        "users": {"imported": 0, "ignored": 0, "errors": []},
        "user_financial_profiles": {"imported": 0, "ignored": 0, "errors": []},
        "user_categories": {"imported": 0, "ignored": 0, "errors": []},
        "cards": {"imported": 0, "ignored": 0, "errors": []},
        "goals": {"imported": 0, "ignored": 0, "errors": []},
        "goal_options": {"imported": 0, "ignored": 0, "errors": []},
        "expenses": {"imported": 0, "ignored": 0, "errors": []},
        "expense_installment_payments": {"imported": 0, "ignored": 0, "errors": []},
    }

    # Whitelists and required columns
    column_configs = {
        "users": {
            "required": ["name", "email"],
            "optional": ["last_name"]
        },
        "user_financial_profiles": {
            "required": ["monthly_salary"],
            "optional": ["auto_mark_installments_paid"]
        },
        "user_categories": {
            "required": ["name"],
            "optional": []
        },
        "cards": {
            "required": ["name", "institution"],
            "optional": ["id", "uuid", "limit", "closing_day", "due_day", "color", "icon"]
        },
        "goals": {
            "required": ["name", "target_amount"],
            "optional": ["id", "uuid", "description", "saved_amount", "deadline", "priority", "category", "status"]
        },
        "goal_options": {
            "required": ["name", "estimated_amount"],
            "optional": ["id", "uuid", "goal_id", "goal_name", "description", "reference_link", "status"]
        },
        "expenses": {
            "required": ["name", "type", "purchase_date", "total_amount", "installment_amount", "category"],
            "optional": ["id", "uuid", "card_id", "card_name", "payment_method", "total_installments", "recurrence_months", "notes", "status", "initial_paid_installments"]
        },
        "expense_installment_payments": {
            "required": ["installment_number"],
            "optional": ["id", "uuid", "expense_id", "expense_name", "paid_at"]
        }
    }

    # Load and map helper
    def get_sheet_rows(internal_sheet_name: str) -> List[Dict[str, Any]]:
        ws = None
        for name in wb.sheetnames:
            if SHEET_MAPPING.get(name.strip().lower()) == internal_sheet_name:
                ws = wb[name]
                break
        
        if ws is None:
            if internal_sheet_name in wb.sheetnames:
                ws = wb[internal_sheet_name]
            else:
                return []
                
        if ws.max_row < 2:
            return []
        
        # Read header row, mapped to DB properties
        raw_headers = [str(cell.value or '').strip() for cell in ws[1]]
        headers = []
        for rh in raw_headers:
            mapped_h = COLUMN_MAPPING.get(rh.lower(), rh.lower())
            headers.append(mapped_h)
        
        # Validate columns
        cfg = column_configs[internal_sheet_name]
        missing = [req for req in cfg["required"] if req not in headers]
        
        # Special case lookup flexibilities
        if internal_sheet_name == "goal_options" and "goal_id" in missing and "goal_name" in headers:
            missing.remove("goal_id")
        if internal_sheet_name == "expenses" and "card_id" in missing and "card_name" in headers:
            missing.remove("card_id")
        if internal_sheet_name == "expense_installment_payments" and "expense_id" in missing and "expense_name" in headers:
            missing.remove("expense_id")
            
        if missing:
            raise ValueError(f"A aba '{internal_sheet_name}' está sem colunas obrigatórias: {', '.join(missing)}")
            
        rows = []
        for r_idx in range(2, ws.max_row + 1):
            row_vals = [cell.value for cell in ws[r_idx]]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
                
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if header:
                    row_dict[header] = row_vals[col_idx]
            rows.append(row_dict)
        return rows

    # Prepopulate maps from existing database items
    card_name_map = {c.name.lower(): c.id for c in db.query(Card).filter(Card.user_id == user.id).all()}
    goal_name_map = {g.name.lower(): g.id for g in db.query(Goal).filter(Goal.user_id == user.id).all()}
    expense_name_map = {e.name.lower(): e.id for e in db.query(Expense).filter(Expense.user_id == user.id).all()}

    card_id_map = {}  # original_card_id -> new_card_id
    goal_id_map = {}  # original_goal_id -> new_goal_id
    expense_id_map = {}  # original_expense_id -> new_expense_id

    # 1. Process `users`
    try:
        user_rows = get_sheet_rows("users")
        for row in user_rows:
            name = str(row.get("name") or "").strip()
            last_name = row.get("last_name")
            if last_name is not None:
                last_name = str(last_name).strip()
            
            if name:
                user.name = name
                user.last_name = last_name
                summary["users"]["imported"] += 1
            else:
                summary["users"]["ignored"] += 1
    except Exception as e:
        summary["users"]["errors"].append(str(e))
        raise

    # 2. Process `user_financial_profiles`
    try:
        profile_rows = get_sheet_rows("user_financial_profiles")
        for row in profile_rows:
            salary = parse_decimal(row.get("monthly_salary"))
            auto_mark_val = row.get("auto_mark_installments_paid")
            if isinstance(auto_mark_val, str) and auto_mark_val.strip().lower() in ("sim", "yes", "s", "y"):
                auto_mark = True
            elif isinstance(auto_mark_val, str) and auto_mark_val.strip().lower() in ("não", "nao", "no", "n"):
                auto_mark = False
            else:
                auto_mark = parse_boolean(auto_mark_val)
                
            if salary is not None:
                profile = db.query(UserFinancialProfile).filter(UserFinancialProfile.user_id == user.id).first()
                if not profile:
                    profile = UserFinancialProfile(user_id=user.id)
                    db.add(profile)
                profile.monthly_salary = salary
                profile.auto_mark_installments_paid = auto_mark
                summary["user_financial_profiles"]["imported"] += 1
            else:
                summary["user_financial_profiles"]["ignored"] += 1
    except Exception as e:
        summary["user_financial_profiles"]["errors"].append(str(e))
        raise

    # 3. Process `user_categories`
    try:
        cat_rows = get_sheet_rows("user_categories")
        for row in cat_rows:
            name = str(row.get("name") or "").strip()
            if not name:
                summary["user_categories"]["ignored"] += 1
                continue
            
            # Check if category exists
            existing = db.query(UserCategory).filter(
                UserCategory.user_id == user.id,
                func.lower(UserCategory.name) == name.lower()
            ).first()
            
            if not existing:
                cat = UserCategory(user_id=user.id, name=name)
                db.add(cat)
                summary["user_categories"]["imported"] += 1
            else:
                summary["user_categories"]["ignored"] += 1
    except Exception as e:
        summary["user_categories"]["errors"].append(str(e))
        raise

    db.flush()

    # 4. Process `cards`
    try:
        card_rows = get_sheet_rows("cards")
        for row in card_rows:
            orig_id = parse_int(row.get("id"))
            name = str(row.get("name") or "").strip()
            institution = str(row.get("institution") or "").strip()
            limit = parse_decimal(row.get("limit"))
            closing_day = parse_int(row.get("closing_day"))
            due_day = parse_int(row.get("due_day"))
            color = str(row.get("color") or "#3B82F6").strip()
            icon = str(row.get("icon") or "credit-card").strip()
            
            if not name or not institution:
                summary["cards"]["ignored"] += 1
                continue

            # Look for duplicate card
            existing = db.query(Card).filter(
                Card.user_id == user.id,
                func.lower(Card.name) == name.lower(),
                func.lower(Card.institution) == institution.lower()
            ).first()

            if existing:
                card_name_map[name.lower()] = existing.id
                if orig_id is not None:
                    card_id_map[orig_id] = existing.id
                summary["cards"]["ignored"] += 1
            else:
                new_card = Card(
                    user_id=user.id,
                    name=name,
                    institution=institution,
                    limit=limit,
                    closing_day=closing_day,
                    due_day=due_day,
                    color=color,
                    icon=icon
                )
                db.add(new_card)
                db.flush()
                card_name_map[name.lower()] = new_card.id
                if orig_id is not None:
                    card_id_map[orig_id] = new_card.id
                summary["cards"]["imported"] += 1
    except Exception as e:
        summary["cards"]["errors"].append(str(e))
        raise

    # 5. Process `goals`
    try:
        goal_rows = get_sheet_rows("goals")
        for row in goal_rows:
            orig_id = parse_int(row.get("id"))
            name = str(row.get("name") or "").strip()
            description = row.get("description")
            if description is not None:
                description = str(description).strip()
            target_amount = parse_decimal(row.get("target_amount"))
            saved_amount = parse_decimal(row.get("saved_amount")) or Decimal("0")
            deadline = parse_date(row.get("deadline"))
            
            priority_val = str(row.get("priority") or "medium").strip().lower()
            priority_in_map = {"baixa": "low", "média": "medium", "media": "medium", "alta": "high"}
            priority_val = priority_in_map.get(priority_val, priority_val)
            priority = GoalPriority.MEDIUM
            if priority_val in [p.value for p in GoalPriority]:
                priority = GoalPriority(priority_val)
                
            cat_val = str(row.get("category") or "other").strip().lower()
            cat_in_map = {
                "carro": "car", "viagem": "travel", "reserva": "reserve",
                "dívida": "debt", "divida": "debt", "casa": "house",
                "educação": "education", "educacao": "education", "saúde": "health",
                "saude": "health", "outro": "other", "outros": "other"
            }
            cat_val = cat_in_map.get(cat_val, cat_val)
            category = GoalCategory.OTHER
            if cat_val in [c.value for c in GoalCategory]:
                category = GoalCategory(cat_val)
                
            status_val = str(row.get("status") or "active").strip().lower()
            status_in_map = {"ativa": "active", "concluída": "completed", "concluida": "completed", "cancelada": "cancelled"}
            status_val = status_in_map.get(status_val, status_val)
            status = GoalStatus.ACTIVE
            if status_val in [s.value for s in GoalStatus]:
                status = GoalStatus(status_val)

            if not name or target_amount is None:
                summary["goals"]["ignored"] += 1
                continue

            # Look for duplicate goal
            existing = db.query(Goal).filter(
                Goal.user_id == user.id,
                func.lower(Goal.name) == name.lower(),
                Goal.target_amount == target_amount
            ).first()

            if existing:
                goal_name_map[name.lower()] = existing.id
                if orig_id is not None:
                    goal_id_map[orig_id] = existing.id
                summary["goals"]["ignored"] += 1
            else:
                new_goal = Goal(
                    user_id=user.id,
                    name=name,
                    description=description,
                    target_amount=target_amount,
                    saved_amount=saved_amount,
                    deadline=deadline,
                    priority=priority,
                    category=category,
                    status=status
                )
                db.add(new_goal)
                db.flush()
                goal_name_map[name.lower()] = new_goal.id
                if orig_id is not None:
                    goal_id_map[orig_id] = new_goal.id
                summary["goals"]["imported"] += 1
    except Exception as e:
        summary["goals"]["errors"].append(str(e))
        raise

    # 6. Process `goal_options`
    try:
        opt_rows = get_sheet_rows("goal_options")
        for row in opt_rows:
            new_goal_id = None
            orig_goal_id = parse_int(row.get("goal_id"))
            if orig_goal_id is not None:
                new_goal_id = goal_id_map.get(orig_goal_id)
            
            if new_goal_id is None:
                goal_name = str(row.get("goal_name") or "").strip()
                if goal_name:
                    new_goal_id = goal_name_map.get(goal_name.lower())
                    
            if not new_goal_id:
                summary["goal_options"]["ignored"] += 1
                continue

            name = str(row.get("name") or "").strip()
            est_amount = parse_decimal(row.get("estimated_amount"))
            description = row.get("description")
            if description is not None:
                description = str(description).strip()
            ref_link = row.get("reference_link")
            if ref_link is not None:
                ref_link = str(ref_link).strip()

            status_val = str(row.get("status") or "analyzing").strip().lower()
            status_in_map = {"em análise": "analyzing", "em analise": "analyzing", "escolhida": "chosen", "descartada": "discarded"}
            status_val = status_in_map.get(status_val, status_val)
            status = GoalOptionStatus.ANALYZING
            if status_val in [s.value for s in GoalOptionStatus]:
                status = GoalOptionStatus(status_val)

            if not name or est_amount is None:
                summary["goal_options"]["ignored"] += 1
                continue

            # Check duplicate goal option
            existing = db.query(GoalOption).filter(
                GoalOption.goal_id == new_goal_id,
                func.lower(GoalOption.name) == name.lower()
            ).first()

            if existing:
                summary["goal_options"]["ignored"] += 1
            else:
                new_opt = GoalOption(
                    goal_id=new_goal_id,
                    name=name,
                    estimated_amount=est_amount,
                    description=description,
                    reference_link=ref_link,
                    status=status
                )
                db.add(new_opt)
                summary["goal_options"]["imported"] += 1
    except Exception as e:
        summary["goal_options"]["errors"].append(str(e))
        raise

    # 7. Process `expenses`
    try:
        expense_rows = get_sheet_rows("expenses")
        for row in expense_rows:
            orig_id = parse_int(row.get("id"))
            
            # Resolve Card ID
            new_card_id = None
            orig_card_id = parse_int(row.get("card_id"))
            if orig_card_id is not None:
                new_card_id = card_id_map.get(orig_card_id)
            
            if new_card_id is None:
                card_name = str(row.get("card_name") or "").strip()
                if card_name:
                    new_card_id = card_name_map.get(card_name.lower())

            name = str(row.get("name") or "").strip()
            total_amount = parse_decimal(row.get("total_amount"))
            installment_amount = parse_decimal(row.get("installment_amount"))
            
            category_val = str(row.get("category") or "").strip()
            category = CATEGORY_MAP.get(category_val.lower(), category_val)
            
            purchase_date = parse_date(row.get("purchase_date"))
            
            pm_val = str(row.get("payment_method") or "credit_card").strip().lower()
            payment_method = PAYMENT_METHOD_MAP.get(pm_val, PaymentMethod.CREDIT_CARD)

            type_val = str(row.get("type") or "one_time").strip().lower()
            expense_type = TYPE_MAP.get(type_val, ExpenseType.ONE_TIME)

            status_val = str(row.get("status") or "active").strip().lower()
            status = STATUS_MAP.get(status_val, ExpenseStatus.ACTIVE)

            total_installments = parse_int(row.get("total_installments")) or 1
            recurrence_months = parse_int(row.get("recurrence_months"))
            initial_paid_installments = parse_int(row.get("initial_paid_installments")) or 0
            notes = row.get("notes")
            if notes is not None:
                notes = str(notes).strip()

            if not name or total_amount is None or installment_amount is None or not category or purchase_date is None:
                summary["expenses"]["ignored"] += 1
                continue

            # Look for duplicate expense
            existing = db.query(Expense).filter(
                Expense.user_id == user.id,
                func.lower(Expense.name) == name.lower(),
                Expense.purchase_date == purchase_date,
                Expense.total_amount == total_amount
            ).first()

            if existing:
                expense_name_map[name.lower()] = existing.id
                if orig_id is not None:
                    expense_id_map[orig_id] = existing.id
                summary["expenses"]["ignored"] += 1
            else:
                new_exp = Expense(
                    user_id=user.id,
                    card_id=new_card_id,
                    payment_method=payment_method,
                    name=name,
                    type=expense_type,
                    purchase_date=purchase_date,
                    total_amount=total_amount,
                    installment_amount=installment_amount,
                    total_installments=total_installments,
                    recurrence_months=recurrence_months,
                    category=category,
                    notes=notes,
                    status=status
                )
                db.add(new_exp)
                db.flush()
                if initial_paid_installments > 0:
                    from app.services.expense_payments import sync_manual_paid_installments
                    try:
                        sync_manual_paid_installments(db, new_exp, initial_paid_installments)
                    except ValueError:
                        pass
                expense_name_map[name.lower()] = new_exp.id
                if orig_id is not None:
                    expense_id_map[orig_id] = new_exp.id
                summary["expenses"]["imported"] += 1
    except Exception as e:
        summary["expenses"]["errors"].append(str(e))
        raise

    # 8. Process `expense_installment_payments`
    try:
        payment_rows = get_sheet_rows("expense_installment_payments")
        for row in payment_rows:
            new_expense_id = None
            orig_expense_id = parse_int(row.get("expense_id"))
            if orig_expense_id is not None:
                new_expense_id = expense_id_map.get(orig_expense_id)
                
            if new_expense_id is None:
                expense_name = str(row.get("expense_name") or "").strip()
                if expense_name:
                    new_expense_id = expense_name_map.get(expense_name.lower())
                    
            if not new_expense_id:
                summary["expense_installment_payments"]["ignored"] += 1
                continue

            inst_num = parse_int(row.get("installment_number"))
            paid_at = parse_datetime(row.get("paid_at")) or datetime.now(timezone.utc)

            if inst_num is None:
                summary["expense_installment_payments"]["ignored"] += 1
                continue

            # Check duplicate payment
            existing = db.query(ExpenseInstallmentPayment).filter(
                ExpenseInstallmentPayment.expense_id == new_expense_id,
                ExpenseInstallmentPayment.installment_number == inst_num
            ).first()

            if existing:
                summary["expense_installment_payments"]["ignored"] += 1
            else:
                new_pay = ExpenseInstallmentPayment(
                    expense_id=new_expense_id,
                    installment_number=inst_num,
                    paid_at=paid_at
                )
                db.add(new_pay)
                summary["expense_installment_payments"]["imported"] += 1
    except Exception as e:
        summary["expense_installment_payments"]["errors"].append(str(e))
        raise

    return summary


def generate_expenses_template(user: User, db: Session) -> io.BytesIO:
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # slate-800
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=11, bold=True)
    data_font = Font(name="Arial", size=10)
    alignment_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Sheet: Instruções
    ws_inst = wb.create_sheet(title="Instruções")
    ws_inst.views.sheetView[0].showGridLines = True
    
    instructions = [
        ["Instruções para Preenchimento do Modelo de Despesas"],
        [""],
        ["1. Não altere os nomes das colunas na primeira linha da aba 'Despesas'."],
        ["2. A aba 'Despesas' é onde você deve cadastrar suas despesas."],
        ["3. Formato de Data na coluna 'Data Compra': AAAA-MM-DD (exemplo: 2026-06-30)."],
        ["4. Formato de Valores (Valor Total e Valor Parcela): Use números (exemplo: 150.00 ou 1500)."],
        ["5. Coluna 'Tipo': Selecione 'Única' ou 'Recorrente' no menu suspenso."],
        ["6. Coluna 'Método Pagamento': Selecione 'Cartão de crédito', 'Cartão de débito', 'Pix' ou 'Boleto' no menu suspenso."],
        ["7. Coluna 'Cartão': Selecione o cartão desejado a partir do menu suspenso (que exibe os cartões da aba 'Cartões Disponíveis')."],
        ["8. Coluna 'Categoria': Selecione uma das opções padrão ou digite uma categoria customizada."],
        ["9. Coluna 'Status': Selecione 'Ativo', 'Pago' ou 'Cancelado'."],
        ["10. Após preencher, salve este arquivo e faça o upload na tela de Dados no aplicativo."]
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        cell = ws_inst.cell(row=row_idx, column=1, value=row_data[0])
        if row_idx == 1:
            cell.font = Font(name="Arial", size=14, bold=True, color="3B82F6")
        elif row_data[0].startswith(tuple(str(i) for i in range(1, 11))):
            cell.font = bold_font
        else:
            cell.font = data_font

    ws_inst.column_dimensions["A"].width = 100

    # 2. Sheet: Despesas (template rows)
    ws_exp = wb.create_sheet(title="Despesas")
    ws_exp.views.sheetView[0].showGridLines = True
    
    headers = [
        "Cartão", "Método Pagamento", "Nome", "Tipo", "Data Compra",
        "Valor Total", "Valor Parcela", "Total Parcelas", "Meses Recorrência",
        "Parcelas Pagas", "Categoria", "Observações", "Status"
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_exp.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_left
        cell.border = thin_border
        
    # Write sample rows
    sample_rows = [
        ["", "Pix", "Exemplo Compra no Pix", "Única", "2026-06-30", 150.00, 150.00, 1, "", 0, "Lazer", "Exemplo de lazer", "Ativo"],
        ["", "Cartão de crédito", "Exemplo Compra Parcelada", "Única", "2026-06-30", 300.00, 100.00, 3, "", 1, "Vestuário", "Exemplo parcelado", "Ativo"]
    ]
    
    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_exp.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = alignment_left
                
    for col in ws_exp.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_exp.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # 3. Sheet: Cartões Disponíveis
    ws_cards = wb.create_sheet(title="Cartões Disponíveis")
    ws_cards.views.sheetView[0].showGridLines = True
    
    cards_headers = ["Nome", "Instituição", "Limite", "Dia Fechamento", "Dia Vencimento", "Cor", "Ícone"]
    for col_idx, header in enumerate(cards_headers, 1):
        cell = ws_cards.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_left
        cell.border = thin_border
        
    cards = db.query(Card).filter(Card.user_id == user.id).all()
    for row_idx, card in enumerate(cards, 2):
        c_name = ws_cards.cell(row=row_idx, column=1, value=card.name)
        c_inst = ws_cards.cell(row=row_idx, column=2, value=card.institution)
        c_limit = ws_cards.cell(row=row_idx, column=3, value=format_cell_value(card.limit))
        c_closing = ws_cards.cell(row=row_idx, column=4, value=format_cell_value(card.closing_day))
        c_due = ws_cards.cell(row=row_idx, column=5, value=format_cell_value(card.due_day))
        c_color = ws_cards.cell(row=row_idx, column=6, value=format_cell_value(card.color))
        c_icon = ws_cards.cell(row=row_idx, column=7, value=format_cell_value(card.icon))
        
        for cell in (c_name, c_inst, c_limit, c_closing, c_due, c_color, c_icon):
            cell.font = data_font
            cell.border = thin_border
            
        c_name.alignment = alignment_left
        c_inst.alignment = alignment_left
        c_limit.alignment = Alignment(horizontal="right", vertical="center")
        c_closing.alignment = Alignment(horizontal="center", vertical="center")
        c_due.alignment = Alignment(horizontal="center", vertical="center")
        c_color.alignment = alignment_left
        c_icon.alignment = alignment_left
        
    for col in ws_cards.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_cards.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Add Dropdown menus (Data Validation) in Excel Template
    dv_cards = DataValidation(type="list", formula1="='Cartões Disponíveis'!$A$2:$A$100", allow_blank=True)
    ws_exp.add_data_validation(dv_cards)
    dv_cards.add("A2:A500")

    dv_pm = DataValidation(type="list", formula1='"Cartão de crédito,Cartão de débito,Pix,Boleto"', allow_blank=True)
    ws_exp.add_data_validation(dv_pm)
    dv_pm.add("B2:B500")

    dv_type = DataValidation(type="list", formula1='"Única,Recorrente"', allow_blank=True)
    ws_exp.add_data_validation(dv_type)
    dv_type.add("D2:D500")

    dv_cat = DataValidation(type="list", formula1='"Alimentação,Moradia,Transporte,Saúde,Educação,Lazer,Vestuário,Assinaturas,Outros"', allow_blank=True)
    ws_exp.add_data_validation(dv_cat)
    dv_cat.add("K2:K500")

    dv_status = DataValidation(type="list", formula1='"Ativo,Pago,Cancelado"', allow_blank=True)
    ws_exp.add_data_validation(dv_status)
    dv_status.add("M2:M500")

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

